import openpyxl as xs
import sys
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

start_time = time.time()

time_load = 2

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--window-size-=1920x1080")
chrome_options.add_argument("log-level=2")


def trackUps(carrier, tn):
    try:
        url = 'https://www.ups.com/track?loc=en_US&tracknum=' + \
            tn + '&requester=WT/trackdetails'
        browser = webdriver.Chrome(chrome_options=chrome_options)
        browser.get(url)
        time.sleep(time_load)
        html = browser.execute_script(
            "return document.documentElement.outerHTML")
        sel_soup = BeautifulSoup(html, 'html.parser')
        date = sel_soup.find(id='stApp_deliveredDate').getText()
        message = 'UPS', date
        return message
    except:
        return 'CHECK: exception (UPS)'


def trackDhl(carrier, tn):
    try:
        url = 'http://www.dhl-usa.com/en/express/tracking.html?AWB=' + tn + '&brand=DHL'
        browser = webdriver.Chrome(chrome_options=chrome_options)
        browser.get(url)
        time.sleep(time_load)
        html = browser.execute_script(
            "return document.documentElement.outerHTML")
        sel_soup = BeautifulSoup(html, 'html.parser')
        spans = sel_soup.find(
            class_='result-summary result-has-pieces').find_all('span')
        date = spans[2].getText()
        months = {
            'January': '01',
            'February': '02',
            'March': '03',
            'April': '04',
            'May': '05',
            'June': '06',
            'July': '07',
            'August': '08',
            'September': '09',
            'October': '10',
            'November': '11',
            'December': '12',
        }

        t, month, day, year, t, t, t = date.split()
        d, t = day.split(',')
        # delivered = sel_soup.find(class_='result-checkpoints show result-has-pieces').tbody.tr.td[1].getText()
        # print
        if month in months:
            message = 'DHL', months[month] + '/' + d + '/' + year
        else:
            message = 'DHL CHECK'
        return message
    except:
        return 'CHECK: exception (DHL)'


def trackFedex(carrier, tn):
    try:
        url = 'https://www.fedex.com/apps/fedextrack/?action=track&tracknumbers=' + \
            tn + '&locale=en_US&cntry_code=us'
        browser = webdriver.Chrome(chrome_options=chrome_options)
        browser.get(url)
        time.sleep(time_load)
        html = browser.execute_script(
            "return document.documentElement.outerHTML")
        sel_soup = BeautifulSoup(html, 'html.parser')
        text = sel_soup.find(
            class_='redesignSnapshotTVC snapshotController_date dest').getText()
        t, date, t, t, t = text.split()
        message = 'FEDEX', date
        return message
    except:
        return 'CHECK: exception (FEDEX)'

file = ''

try:
    wb = xs.load_workbook(file)
    sheet = wb['Sheet1']
except:
    sys.exit('CHECK: error in loading excel')


i = 1
date_list = []
carrierList = ['UPS', 'DHL', 'FEDEX']
while sheet.cell(i, 7).value != None:
    line = sheet.cell(i, 7).value

    carrier, track_number = line.split()

    if carrier.upper() == 'UPS':
        date = trackUps(carrier, track_number)
    if carrier.upper() == 'DHL':
        date = trackDhl(carrier, track_number)
    if carrier.upper() == 'FEDEX':
        date = trackFedex(carrier, track_number)
    elif carrier.upper() not in carrierList:
        date = 'CHECK: carrier not found'

    date_list.append(date)
    if 'CHECK' in date:
        sheet.cell(i, 8).value = date
    else:
        sheet.cell(i, 8).value = date[1]

    i = i + 1


for date in date_list:
    print(date)

print('run time:', (time.time()-start_time))


wb.save(file)
