import openpyxl as xs

curWeekSchedule = 'SCHNOV1824'
curMonthSchedule = 'statsNovValet'

wb = xs.load_workbook()
sheet = wb['Table 1']


class worker:
    def __init__(self, name, shiftsWorked, goodShifts, normalShift, badShifts, daysOff):
        self.name = name
        self.shiftsWorked = shiftsWorked
        self.goodShifts = goodShifts
        self.normalShift = normalShift
        self.badShifts = badShifts
        self.daysOff = daysOff


ROW = 33
COL = 8
names = []
person = []
places = []
days = []


def cleanNames(table, names, ROW, COL):
    for row in range(1, ROW):
        for col in range(1, COL):
            if (table[row][col] not in names) and (
                    table[row][col] != None) and (
                    table[row][col] != 'MON') and (
                    table[row][col] != 'TUE') and (
                    table[row][col] != 'WED') and (
                    table[row][col] != 'THU') and (
                    table[row][col] != 'FRI') and (
                    table[row][col] != 'SAT') and (
                    table[row][col] != 'SUN') and (
                    table[row][col] != 'X'):
                names.append(table[row][col])
    names.sort()


def cleanPlaces(table, places, ROW):
    for row in range(3, ROW):
        if (table[row][0] not in places) and (
                table[row][0] != '3:45 pm') and (
                table[row][0] != '5:00 pm') and (
                table[row][0] != '6:00 pm') and (
                table[row][0] != '4:45 pm') and (
                table[row][0] != '5:30 pm') and (
                table[row][0] != '04:45 pm') and (
                table[row][0] != '8:00 am') and (
                table[row][0] != '8:00 am  (front)') and (
                table[row][0] != '8:00 am (back)') and (
                table[row][0] != '3:00 pm') and (
                table[row][0] != '7: 00 am') and (
                table[row][0] != '*09:00 am 02:00pm*') and (
                table[row][0] != '6:00 am'):
            places.append(table[row][0])


def createDays(table, days, COL):
    for col in range(1, COL):
        days.append(table[3][col])


# change hours in the table to same name place to make the process easier
def modTablePlaces(table, places, ROW):
    for row in range(ROW - 1):
        if table[row][0] in places:
            while table[row][0] not in table[row + 1][0] and table[row + 1][0] not in places:
                table[row + 1][0] = table[row][0]


def roccosShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'ROCCO’S  TACOS'):
            #  roccos good shifts
            if (table[3][cols] == 'TUE' or table[3][cols] == 'FRI' or table[3][cols] == 'SAT'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].goodShifts = person[i].goodShifts + 1
            #  roccos normal shifts
            if (table[3][cols] == 'MON' or table[3][cols] == 'WED' or table[3][cols] == 'THU' or table[3][cols] == 'SUN'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].normalShift = person[i].normalShift + 1


def agaveShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'AGAVE AZUL'):
            #  agave good shifts
            if (table[3][cols] == 'FRI' or table[3][cols] == 'SAT'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].goodShifts = person[i].goodShifts + 1
            #  roccos normal shifts
            if (table[3][cols] == 'TUE' or table[3][cols] == 'THU'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].normalShift = person[i].normalShift + 1


def shulasShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'SHULAS 347'):
            #  shulas bad shifts
            if (table[3][cols] == 'WED' or table[3][cols] == 'THU' or table[3][cols] == 'FRI' or table[3][cols] == 'SAT'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].badShifts = person[i].badShifts + 1


def marlowsShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'MARLOW’S  TAVERN'):
            #  MARLOW’S  TAVERN bad shifts
            if (table[3][cols] == 'FRI' or table[3][cols] == 'SAT'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].badShifts = person[i].badShifts + 1


def jeweetShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'JEWEET CLINIC'):
            #  JEWEET normal shifts
            if (table[3][cols] == 'MON' or table[3][cols] == 'TUE' or table[3][cols] == 'WED' or table[3][cols] == 'THU' or table[3][cols] == 'FRI'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].normalShift = person[i].normalShift + 1


def christnersShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'CHRISTNER’S'):
            #  CHRISTNER’S good shifts
            if (table[3][cols] == 'MON' or table[3][cols] == 'TUE' or table[3][cols] == 'WED' or table[3][cols] == 'THU' or table[3][cols] == 'FRI' or table[3][cols] == 'SAT'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].goodShifts = person[i].goodShifts + 1


def evsShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'EDDIE  V’S'):
            #  EDDIE  V’S good shifts
            if (table[3][cols] == 'MON' or table[3][cols] == 'TUE' or table[3][cols] == 'WED' or table[3][cols] == 'THU' or table[3][cols] == 'FRI' or table[3][cols] == 'SAT' or table[3][cols] == 'SUN'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].goodShifts = person[i].goodShifts + 1


def magruderShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'MAGRUDER'):
            #  MAGRUDER normal shifts
            if (table[3][cols] == 'MON' or table[3][cols] == 'TUE' or table[3][cols] == 'WED' or table[3][cols] == 'THU' or table[3][cols] == 'FRI'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].normalShift = person[i].normalShift + 1


def cdhShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'CDH'):
            #  CDH normal shifts
            if (table[3][cols] == 'MON' or table[3][cols] == 'TUE' or table[3][cols] == 'WED' or table[3][cols] == 'THU' or table[3][cols] == 'FRI'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].normalShift = person[i].normalShift + 1


def kabookiShifts(table, person, places, days, ROW, COL, i, rows, cols):
    if (table[rows][cols] == person[i].name):
        if (table[rows][0] == 'KABOOKI'):
            #  KABOOKI normal shifts
            if (table[3][cols] == 'FRI' or table[3][cols] == 'SAT'):
                person[i].shiftsWorked = person[i].shiftsWorked + 1
                person[i].normalShift = person[i].normalShift + 1


def scanShifts(table, person, places, days, ROW, COL):
    for i in range(len(person)):
        for rows in range(4, ROW):
            for cols in range(1, COL):
#               evaluate shifts of Rocco's tacos
                roccosShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate shifts of agave
                agaveShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate shulas Shifts
                shulasShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate MARLOW’S  TAVERN
                marlowsShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate JEWEET CLINIC
                jeweetShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate CHRISTNER’S
                christnersShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate EDDIE  V’S
                evsShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate MAGRUDER
                magruderShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate CDH
                cdhShifts(table, person, places, days, ROW, COL, i, rows, cols)
#               evaluate KABOOKI
                kabookiShifts(table, person, places, days, ROW, COL, i, rows, cols)


def daysOffCheck(table, person, places, days, ROW, COL):
    for i in range(len(person)):
        for cols in range(1, COL):
            count = 0
            for rows in range(4, ROW):
                if(table[rows][cols] == person[i].name):
                    count = count + 1
            if(count == 0):
                person[i].daysOff = person[i].daysOff + 1



table = [
    [] for i in range(ROW)
]

for row in range(ROW):
    for col in range(COL):
        table[row].insert(col, sheet.cell(row + 1, col + 1).value)

cleanNames(table, names, ROW, COL)
for i in range(len(names)):
    person.append(worker(names[i], 0, 0, 0, 0, 0))

cleanPlaces(table, places, ROW)
createDays(table, days, COL)
modTablePlaces(table, places, ROW)
scanShifts(table, person, places, days, ROW, COL)
daysOffCheck(table, person, places, days, ROW, COL)

# for i in range(len(person)):
#     print(person[i].name + " " + str(person[i].shiftsWorked) + " " + str(person[i].goodShifts) +
#           " " + str(person[i].normalShift) + " " + str(person[i].badShifts) + " " + str(person[i].daysOff))


sheet.cell(2, 10).value = 'workers'
sheet.cell(2, 11).value = 'shiftsWorked'
sheet.cell(2, 12).value = 'goodShifts'
sheet.cell(2, 13).value = 'normalShift'
sheet.cell(2, 14).value = 'badShifts'
sheet.cell(2, 15).value = 'daysOff'

for row in range(3, len(names) + 3):
    sheet.cell(row, 10).value = person[row - 3].name
    sheet.cell(row, 11).value = person[row - 3].shiftsWorked
    sheet.cell(row, 12).value = person[row - 3].goodShifts
    sheet.cell(row, 13).value = person[row - 3].normalShift
    sheet.cell(row, 14).value = person[row - 3].badShifts
    sheet.cell(row, 15).value = person[row - 3].daysOff


wb.save()

# monthWB = xs.load_workbook()
# monthSheet = monthWB['Sheet1']
# weeks = []
# n = monthSheet.cell(1, 1).value
#
# for i in range(n):
#     weeks.append(monthSheet.cell(n+1, 1).value)
#
# if(curWeekSchedule not in weeks):
#     monthSheet.cell(2, 3).value = 'workers'
#     monthSheet.cell(2, 4).value = 'shiftsWorked'
#     monthSheet.cell(2, 5).value = 'goodShifts'
#     monthSheet.cell(2, 6).value = 'normalShift'
#     monthSheet.cell(2, 7).value = 'badShifts'
#     monthSheet.cell(2, 8).value = 'daysOff'
#
#     # exact numbers
#     for row in range(3, 23):
#         monthSheet.cell(row, 3).value = person[row - 3].name
#         monthSheet.cell(row, 4).value = monthSheet.cell(row, 4).value + person[row - 3].shiftsWorked
#         monthSheet.cell(row, 5).value = monthSheet.cell(row, 5).value + person[row - 3].goodShifts
#         monthSheet.cell(row, 6).value = monthSheet.cell(row, 6).value + person[row - 3].normalShift
#         monthSheet.cell(row, 7).value = monthSheet.cell(row, 7).value + person[row - 3].badShifts
#         monthSheet.cell(row, 8).value = monthSheet.cell(row, 8).value + person[row - 3].daysOff
#
#     monthSheet.cell(2, 11).value = 'workers'
#     monthSheet.cell(2, 12).value = 'shiftsWorked'
#     monthSheet.cell(2, 13).value = 'goodShifts'
#     monthSheet.cell(2, 14).value = 'normalShift'
#     monthSheet.cell(2, 15).value = 'badShifts'
#     monthSheet.cell(2, 16).value = 'daysOff'
#
#     # percentages
#     for row in range(3, 23):
#         monthSheet.cell(row, 11).value = person[row - 3].name
#         monthSheet.cell(row, 12).value = 100
#         monthSheet.cell(row, 13).value = round((monthSheet.cell(row, 5).value * 100) / monthSheet.cell(row, 4).value, 1)
#         monthSheet.cell(row, 14).value = round((monthSheet.cell(row, 6).value * 100) / monthSheet.cell(row, 4).value, 1)
#         monthSheet.cell(row, 15).value = round((monthSheet.cell(row, 7).value * 100) / monthSheet.cell(row, 4).value, 1)
#         monthSheet.cell(row, 16).value = round(((monthSheet.cell(row, 8).value * 100) / 7), 1)
#
#
#     monthSheet.cell(1, 1).value = monthSheet.cell(1, 1).value + 1
#     monthSheet.cell(monthSheet.cell(1, 1).value + 1, 1).value = curWeekSchedule
#
#     monthWB.save()
#
# allTimeWB = xs.load_workbook()
# allTimeSheet = allTimeWB['Sheet1']
#
# weeks2 = []
# n2 = allTimeSheet.cell(1, 1).value
#
# for i in range(n2):
#     weeks2.append(allTimeSheet.cell(n2+1, 1).value)
#
# if(curWeekSchedule not in weeks2):
#     allTimeSheet.cell(2, 3).value = 'workers'
#     allTimeSheet.cell(2, 4).value = 'shiftsWorked'
#     allTimeSheet.cell(2, 5).value = 'goodShifts'
#     allTimeSheet.cell(2, 6).value = 'normalShift'
#     allTimeSheet.cell(2, 7).value = 'badShifts'
#     allTimeSheet.cell(2, 8).value = 'daysOff'
#
#     # exact numbers
#     for row in range(3, 23):
#         allTimeSheet.cell(row, 3).value = person[row - 3].name
#         allTimeSheet.cell(row, 4).value = allTimeSheet.cell(row, 4).value + sheet.cell(row, 11).value
#         allTimeSheet.cell(row, 5).value = allTimeSheet.cell(row, 5).value + sheet.cell(row, 12).value
#         allTimeSheet.cell(row, 6).value = allTimeSheet.cell(row, 6).value + sheet.cell(row, 13).value
#         allTimeSheet.cell(row, 7).value = allTimeSheet.cell(row, 7).value + sheet.cell(row, 14).value
#         allTimeSheet.cell(row, 8).value = allTimeSheet.cell(row, 8).value + sheet.cell(row, 15).value
#
#     allTimeSheet.cell(2, 11).value = 'workers'
#     allTimeSheet.cell(2, 12).value = 'shiftsWorked'
#     allTimeSheet.cell(2, 13).value = 'goodShifts'
#     allTimeSheet.cell(2, 14).value = 'normalShift'
#     allTimeSheet.cell(2, 15).value = 'badShifts'
#     allTimeSheet.cell(2, 16).value = 'daysOff'
#
#     # percentages
#     for row in range(3, 23):
#         allTimeSheet.cell(row, 11).value = person[row - 3].name
#         allTimeSheet.cell(row, 12).value = 100
#         allTimeSheet.cell(row, 13).value = round((allTimeSheet.cell(row, 5).value * 100) / allTimeSheet.cell(row, 4).value, 1)
#         allTimeSheet.cell(row, 14).value = round((allTimeSheet.cell(row, 6).value * 100) / allTimeSheet.cell(row, 4).value, 1)
#         allTimeSheet.cell(row, 15).value = round((allTimeSheet.cell(row, 7).value * 100) / allTimeSheet.cell(row, 4).value, 1)
#         allTimeSheet.cell(row, 16).value = round(((allTimeSheet.cell(row, 8).value * 100) / 7), 1)
#
#
#     allTimeSheet.cell(1, 1).value = allTimeSheet.cell(1, 1).value + 1
#     allTimeSheet.cell(allTimeSheet.cell(1, 1).value + 1, 1).value = curWeekSchedule
#
#     allTimeWB.save()
